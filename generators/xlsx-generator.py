from __future__ import annotations

import re
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── translations ──────────────────────────────────────────────────────────────

TRANSLATIONS: dict[str, dict[str, str]] = {
    "en": {
        # summary metadata
        "meta_period":      "Financial Period",
        "meta_start":       "Period Start",
        "meta_end":         "Period End",
        "meta_generated":   "Generated",
        # summary table headers
        "sum_company":      "Company",
        "sum_sites":        "Sites",
        "sum_invoices":     "Invoices",
        "sum_heat_cost":    "Heat Cost (£)",
        "sum_cap_charge":   "Capacity Charge (£)",
        "sum_subtotal":     "Subtotal (£)",
        "sum_vat":          "VAT (£)",
        "sum_total":        "Total Due (£)",
        "sum_grand":        "TOTAL",
        # detail column headers
        "col_invoice_no":   "Invoice No",
        "col_company":      "Company",
        "col_site":         "Site",
        "col_city":         "City",
        "col_postcode":     "Postcode",
        "col_meter_id":     "Meter ID",
        "col_period":       "Billing Period",
        "col_period_start": "Period Start",
        "col_period_end":   "Period End",
        "col_issue_date":   "Issue Date",
        "col_due_date":     "Due Date",
        "col_prev_read":    "Prev Reading (kWh)",
        "col_curr_read":    "Curr Reading (kWh)",
        "col_consumption":  "Consumption (kWh)",
        "col_unit_price":   "Unit Price (£/kWh)",
        "col_heat_cost":    "Heat Cost (£)",
        "col_capacity":     "Capacity (kW)",
        "col_cap_rate":     "Cap. Rate (£/kW/mo)",
        "col_supplier_ef": "Supplier EF (kg CO\u2082e/kWh)",
        "col_cap_charge":   "Capacity Charge (£)",
        "col_subtotal":     "Subtotal (£)",
        "col_vat":          "VAT (£)",
        "col_total":        "Total (£)",
    },
    "fr": {
        "meta_period":      "Période financière",
        "meta_start":       "Début de période",
        "meta_end":         "Fin de période",
        "meta_generated":   "Généré",
        "sum_company":      "Entreprise",
        "sum_sites":        "Sites",
        "sum_invoices":     "Factures",
        "sum_heat_cost":    "Coût thermique (£)",
        "sum_cap_charge":   "Frais de capacité (£)",
        "sum_subtotal":     "Sous-total (£)",
        "sum_vat":          "TVA (£)",
        "sum_total":        "Total dû (£)",
        "sum_grand":        "TOTAL",
        "col_invoice_no":   "N° de facture",
        "col_company":      "Entreprise",
        "col_site":         "Site",
        "col_city":         "Ville",
        "col_postcode":     "Code postal",
        "col_meter_id":     "ID compteur",
        "col_period":       "Période de facturation",
        "col_period_start": "Début de période",
        "col_period_end":   "Fin de période",
        "col_issue_date":   "Date d'émission",
        "col_due_date":     "Date d'échéance",
        "col_prev_read":    "Relevé préc. (kWh)",
        "col_curr_read":    "Relevé actuel (kWh)",
        "col_consumption":  "Consommation (kWh)",
        "col_unit_price":   "Prix unit. (£/kWh)",
        "col_heat_cost":    "Coût thermique (£)",
        "col_capacity":     "Capacité (kW)",
        "col_cap_rate":     "Taux cap. (£/kW/mois)",
        "col_supplier_ef": "FE fournisseur (kg CO\u2082e/kWh)",
        "col_cap_charge":   "Frais de cap. (£)",
        "col_subtotal":     "Sous-total (£)",
        "col_vat":          "TVA (£)",
        "col_total":        "Total (£)",
    },
    "de": {
        "meta_period":      "Finanzzeitraum",
        "meta_start":       "Zeitraum Beginn",
        "meta_end":         "Zeitraum Ende",
        "meta_generated":   "Erstellt",
        "sum_company":      "Unternehmen",
        "sum_sites":        "Standorte",
        "sum_invoices":     "Rechnungen",
        "sum_heat_cost":    "Wärmekosten (£)",
        "sum_cap_charge":   "Leistungsgebühr (£)",
        "sum_subtotal":     "Zwischensumme (£)",
        "sum_vat":          "MwSt. (£)",
        "sum_total":        "Gesamtbetrag (£)",
        "sum_grand":        "GESAMT",
        "col_invoice_no":   "Rechnungsnr.",
        "col_company":      "Unternehmen",
        "col_site":         "Standort",
        "col_city":         "Stadt",
        "col_postcode":     "Postleitzahl",
        "col_meter_id":     "Zähler-ID",
        "col_period":       "Abrechnungszeitraum",
        "col_period_start": "Zeitraum Beginn",
        "col_period_end":   "Zeitraum Ende",
        "col_issue_date":   "Ausstellungsdatum",
        "col_due_date":     "Fälligkeitsdatum",
        "col_prev_read":    "Vorh. Stand (kWh)",
        "col_curr_read":    "Akt. Stand (kWh)",
        "col_consumption":  "Verbrauch (kWh)",
        "col_unit_price":   "Einheitspreis (£/kWh)",
        "col_heat_cost":    "Wärmekosten (£)",
        "col_capacity":     "Leistung (kW)",
        "col_cap_rate":     "Leistungssatz (£/kW/Mo)",
        "col_supplier_ef": "EF Lieferant (kg CO\u2082e/kWh)",
        "col_cap_charge":   "Leistungsgebühr (£)",
        "col_subtotal":     "Zwischensumme (£)",
        "col_vat":          "MwSt. (£)",
        "col_total":        "Gesamt (£)",
    },
    "nl": {
        "meta_period":      "Financiële periode",
        "meta_start":       "Periode begin",
        "meta_end":         "Periode einde",
        "meta_generated":   "Gegenereerd",
        "sum_company":      "Bedrijf",
        "sum_sites":        "Locaties",
        "sum_invoices":     "Facturen",
        "sum_heat_cost":    "Warmtekosten (£)",
        "sum_cap_charge":   "Vermogenstoeslag (£)",
        "sum_subtotal":     "Subtotaal (£)",
        "sum_vat":          "BTW (£)",
        "sum_total":        "Totaal (£)",
        "sum_grand":        "TOTAAL",
        "col_invoice_no":   "Factuurnr.",
        "col_company":      "Bedrijf",
        "col_site":         "Locatie",
        "col_city":         "Stad",
        "col_postcode":     "Postcode",
        "col_meter_id":     "Meter-ID",
        "col_period":       "Facturatieperiode",
        "col_period_start": "Periode begin",
        "col_period_end":   "Periode einde",
        "col_issue_date":   "Uitgiftedatum",
        "col_due_date":     "Vervaldatum",
        "col_prev_read":    "Vorige stand (kWh)",
        "col_curr_read":    "Huidige stand (kWh)",
        "col_consumption":  "Verbruik (kWh)",
        "col_unit_price":   "Eenheidsprijs (£/kWh)",
        "col_heat_cost":    "Warmtekosten (£)",
        "col_capacity":     "Vermogen (kW)",
        "col_cap_rate":     "Vermogenstarief (£/kW/mnd)",
        "col_supplier_ef": "EF leverancier (kg CO\u2082e/kWh)",
        "col_cap_charge":   "Vermogenstoeslag (£)",
        "col_subtotal":     "Subtotaal (£)",
        "col_vat":          "BTW (£)",
        "col_total":        "Totaal (£)",
    },
}


# ── palette ───────────────────────────────────────────────────────────────────
_WHITE = "FFFFFF"
_DARK = "1F2328"
_MID = "5A6066"
_ALT = "F7F9FC"
_SOFT = "DCEBF5"
_BORDER = "C9CDD2"

# ── style primitives ──────────────────────────────────────────────────────────

def _font(bold: bool = False, color: str = _DARK, size: int = 10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _border() -> Border:
    side = Side(style="thin", color=_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _align(horizontal: str = "left") -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center")


def _header_cell(cell, text: str, accent: str) -> None:
    cell.value = text
    cell.font = _font(bold=True, color=_WHITE)
    cell.fill = _fill(accent)
    cell.border = _border()
    cell.alignment = _align("center")


def _data_cell(cell, value: Any, fmt: str | None = None, alt: bool = False, bold: bool = False) -> None:
    cell.value = value
    cell.font = _font(bold=bold)
    cell.fill = _fill(_ALT if alt else _WHITE)
    cell.border = _border()
    cell.alignment = _align("right" if isinstance(value, (int, float, Decimal)) else "left")
    if fmt:
        cell.number_format = fmt


# ── detail sheet column spec: (header_key, width, number_format) ─────────────
_DETAIL_COL_SPEC: list[tuple[str, int, str | None]] = [
    ("col_invoice_no",   20, None),
    ("col_company",      30, None),
    ("col_site",         20, None),
    ("col_city",         14, None),
    ("col_postcode",     11, None),
    ("col_meter_id",     22, None),
    ("col_period",       18, None),
    ("col_period_start", 14, "DD MMM YYYY"),
    ("col_period_end",   14, "DD MMM YYYY"),
    ("col_issue_date",   14, "DD MMM YYYY"),
    ("col_due_date",     14, "DD MMM YYYY"),
    ("col_prev_read",    20, "#,##0"),
    ("col_curr_read",    20, "#,##0"),
    ("col_consumption",  20, "#,##0"),
    ("col_unit_price",   20, "0.000"),
    ("col_heat_cost",    16, "#,##0.00"),
    ("col_capacity",     14, "#,##0"),
    ("col_cap_rate",     20, "0.00"),
    ("col_supplier_ef",  24, "0.0000"),
    ("col_cap_charge",   20, "#,##0.00"),
    ("col_subtotal",     16, "#,##0.00"),
    ("col_vat",          13, "#,##0.00"),
    ("col_total",        16, "#,##0.00"),
]


def _detail_cols(strings: dict) -> list[tuple[str, int, str | None]]:
    return [(strings[key], width, fmt) for key, width, fmt in _DETAIL_COL_SPEC]


# ── helpers ───────────────────────────────────────────────────────────────────

def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Return a valid Excel sheet name (max 31 chars, no forbidden characters)."""
    sanitised = re.sub(r'[\\/?*\[\]:]', '', name).strip()
    return (sanitised or "Company")[:max_len]


def _sections_by_company(sections: list[dict]) -> dict[str, list[dict]]:
    """Group sections by company label, preserving insertion order."""
    grouped: dict[str, list[dict]] = {}
    for section in sections:
        label = section["company"]["label"]
        grouped.setdefault(label, []).append(section)
    return grouped


# ── public API ────────────────────────────────────────────────────────────────

def generate_xlsx(
    config: dict,
    sections: list[dict],
    blank_fields: set[str] | None = None,
    split_by_company: bool = False,
) -> bytes:
    """Build a styled XLSX workbook from billing sections and return bytes.

    blank_fields: record field names whose cells should be left empty (QA testing).
    split_by_company: when True, create one detail sheet per company instead of
                      a single combined "Billing Detail" sheet.
    """
    lang = config["document"].get("language", "en")
    strings = TRANSLATIONS.get(lang, TRANSLATIONS["en"])

    default_accent = sections[0]["company"]["accent"].lstrip("#") if sections else "1E5B88"
    omit = blank_fields or set()

    wb = openpyxl.Workbook()
    wb.active.title = "Summary"
    _build_summary_sheet(wb.active, config, sections, default_accent, strings)

    if split_by_company:
        for label, co_sections in _sections_by_company(sections).items():
            accent = co_sections[0]["company"]["accent"].lstrip("#")
            sheet_name = _safe_sheet_name(label)
            _build_detail_sheet(wb.create_sheet(sheet_name), co_sections, accent, omit, strings)
    else:
        _build_detail_sheet(wb.create_sheet("Billing Detail"), sections, default_accent, omit, strings)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── summary sheet ─────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, config: dict, sections: list[dict], accent: str, strings: dict) -> None:
    # Title banner
    ws.merge_cells("A1:H1")
    ws["A1"].value = config["document"]["title"]
    ws["A1"].font = _font(bold=True, color=_WHITE, size=14)
    ws["A1"].fill = _fill(accent)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 30

    # Metadata
    meta = [
        (strings["meta_period"],    config["financial_period"]["label"]),
        (strings["meta_start"],     config["financial_period"]["start_date"].strftime("%d %b %Y")),
        (strings["meta_end"],       config["financial_period"]["end_date"].strftime("%d %b %Y")),
        (strings["meta_generated"], datetime.now().strftime("%d %b %Y %H:%M")),
    ]
    for offset, (label, value) in enumerate(meta):
        row = 2 + offset
        ws.cell(row=row, column=1, value=label).font = _font(bold=True, color=_MID)
        ws.cell(row=row, column=2, value=value).font = _font()

    # Company summary table
    tbl_row = 2 + len(meta) + 1
    summary_headers = [
        strings["sum_company"], strings["sum_sites"],    strings["sum_invoices"],
        strings["sum_heat_cost"], strings["sum_cap_charge"],
        strings["sum_subtotal"], strings["sum_vat"],     strings["sum_total"],
    ]
    for col, h in enumerate(summary_headers, start=1):
        _header_cell(ws.cell(row=tbl_row, column=col), h, accent)

    # Aggregate per company
    totals: dict[str, dict] = {}
    for section in sections:
        key = section["company"]["label"]
        if key not in totals:
            totals[key] = {
                "sites": set(),
                "invoices": 0,
                "heat_cost": Decimal("0"),
                "capacity_charge": Decimal("0"),
                "subtotal": Decimal("0"),
                "vat": Decimal("0"),
                "total": Decimal("0"),
            }
        t = totals[key]
        t["sites"].add(section["site"]["label"])
        for rec in section["records"]:
            t["invoices"] += 1
            t["heat_cost"] += rec["heat_cost"]
            t["capacity_charge"] += rec["capacity_charge"]
            t["subtotal"] += rec["subtotal"]
            t["vat"] += rec["vat"]
            t["total"] += rec["total"]

    money_fmt = "#,##0.00"
    for i, (company, t) in enumerate(totals.items()):
        row = tbl_row + 1 + i
        alt = i % 2 == 1
        row_values = [
            (company,               None),
            (len(t["sites"]),       "#,##0"),
            (t["invoices"],         "#,##0"),
            (float(t["heat_cost"]), money_fmt),
            (float(t["capacity_charge"]), money_fmt),
            (float(t["subtotal"]),  money_fmt),
            (float(t["vat"]),       money_fmt),
            (float(t["total"]),     money_fmt),
        ]
        for col, (val, fmt) in enumerate(row_values, start=1):
            _data_cell(ws.cell(row=row, column=col), val, fmt=fmt, alt=alt)

    # Grand total row
    grand_row = tbl_row + 1 + len(totals)
    grand_values = [
        (strings["sum_grand"], None),
        ("",      None),
        (sum(t["invoices"] for t in totals.values()), "#,##0"),
        (float(sum(t["heat_cost"]        for t in totals.values())), money_fmt),
        (float(sum(t["capacity_charge"]  for t in totals.values())), money_fmt),
        (float(sum(t["subtotal"]         for t in totals.values())), money_fmt),
        (float(sum(t["vat"]              for t in totals.values())), money_fmt),
        (float(sum(t["total"]            for t in totals.values())), money_fmt),
    ]
    for col, (val, fmt) in enumerate(grand_values, start=1):
        c = ws.cell(row=grand_row, column=col, value=val)
        c.font = _font(bold=True)
        c.fill = _fill(_SOFT)
        c.border = _border()
        c.alignment = _align("right" if col > 1 else "left")
        if fmt:
            c.number_format = fmt

    # Column widths
    for col, width in enumerate([30, 8, 10, 20, 22, 18, 14, 20], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"


# ── detail sheet ──────────────────────────────────────────────────────────────

def _build_detail_sheet(ws, sections: list[dict], accent: str, blank_fields: set[str], strings: dict) -> None:
    # Headers
    for col, (label, width, _) in enumerate(_detail_cols(strings), start=1):
        _header_cell(ws.cell(row=1, column=col), label, accent)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Each entry: (value, record_field_name_or_None, fmt)
    # record_field_name is checked against blank_fields to omit the cell value.
    row = 2
    for section in sections:
        company = section["company"]
        site = section["site"]
        for rec in section["records"]:
            alt = row % 2 == 0
            row_data: list[tuple[Any, str | None, str | None]] = [
                (rec["invoice_no"],          "invoice_no",    None),
                (company["label"],           "company_label", None),
                (site["label"],              "site_label",    None),
                (rec["city"],                "city",          None),
                (rec["postcode"],            "postcode",      None),
                (rec["meter_id"],            "meter_id",      None),
                (rec["billing_period_label"],"period_label",  None),
                (rec["period_start"],        None,            "DD MMM YYYY"),
                (rec["period_end"],          None,            "DD MMM YYYY"),
                (rec["issue_date"],          None,            "DD MMM YYYY"),
                (rec["due_date"],            None,            "DD MMM YYYY"),
                (rec["prev_read"],           "prev_read",     "#,##0"),
                (rec["curr_read"],           "curr_read",     "#,##0"),
                (rec["consumption"],         "consumption",   "#,##0"),
                (float(rec["unit_price"]),   "unit_price",    "0.000"),
                (float(rec["heat_cost"]),    "heat_cost",     "#,##0.00"),
                (rec["capacity_kw"],         "capacity_kw",   "#,##0"),
                (float(rec["capacity_rate"]),"capacity_rate", "0.00"),
                (float(rec["supplier_ef"]),  "supplier_ef",   "0.0000"),
                (float(rec["capacity_charge"]), "capacity_charge", "#,##0.00"),
                (float(rec["subtotal"]),     "subtotal",      "#,##0.00"),
                (float(rec["vat"]),          "vat",           "#,##0.00"),
                (float(rec["total"]),        "total",         "#,##0.00"),
            ]
            for col, (value, field_name, fmt) in enumerate(row_data, start=1):
                cell_value = None if (field_name and field_name in blank_fields) else value
                _data_cell(ws.cell(row=row, column=col), cell_value, fmt=fmt, alt=alt)
            ws.row_dimensions[row].height = 18
            row += 1
