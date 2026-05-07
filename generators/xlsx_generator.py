from __future__ import annotations

import re
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.currency import currency_code
from utils.distractor_fields import resolve_tabular_value

# ── translations ──────────────────────────────────────────────────────────────

TRANSLATIONS: dict[str, dict[str, str]] = {
    "en": {
        # summary metadata
        "meta_period":      "Financial Period",
        "meta_start":       "Period Start",
        "meta_end":         "Period End",
        "meta_generated":   "Generated",
        # summary table headers
        "sum_company":      "Company",
        "sum_currency":     "Currency",
        "sum_sites":        "Sites",
        "sum_invoices":     "Invoices",
        "sum_heat_cost":    "Heat Cost (£)",
        "sum_cap_charge":   "Capacity Charge (£)",
        "sum_subtotal":     "Subtotal (£)",
        "sum_vat":          "VAT (£)",
        "sum_total":        "Total Due (£)",
        "sum_grand":        "TOTAL",
        # detail column headers
        "col_invoice_no":   "Invoice No",
        "col_company":      "Company",
        "col_currency":     "Currency",
        "col_site":         "Site",
        "col_city":         "City",
        "col_postcode":     "Postcode",
        "col_meter_id":     "Meter ID",
        "col_period":       "Billing Period",
        "col_period_start": "Period Start",
        "col_period_end":   "Period End",
        "col_issue_date":   "Issue Date",
        "col_due_date":     "Due Date",
        "col_prev_read":    "Prev Reading (kWh)",
        "col_curr_read":    "Curr Reading (kWh)",
        "col_consumption":  "Consumption (kWh)",
        "col_unit_price":   "Unit Price (£/kWh)",
        "col_heat_cost":    "Heat Cost (£)",
        "col_capacity":     "Capacity (kW)",
        "col_cap_rate":     "Cap. Rate (£/kW/mo)",
        "col_supplier_ef": "Supplier EF (kg CO\u2082e/kWh)",
        "col_cap_charge":   "Capacity Charge (£)",
        "col_subtotal":     "Subtotal (£)",
        "col_vat":          "VAT (£)",
        "col_total":        "Total (£)",
    },
    "fr": {
        "meta_period":      "Période financière",
        "meta_start":       "Début de période",
        "meta_end":         "Fin de période",
        "meta_generated":   "Généré",
        "sum_company":      "Entreprise",
        "sum_currency":     "Devise",
        "sum_sites":        "Sites",
        "sum_invoices":     "Factures",
        "sum_heat_cost":    "Coût thermique (£)",
        "sum_cap_charge":   "Frais de capacité (£)",
        "sum_subtotal":     "Sous-total (£)",
        "sum_vat":          "TVA (£)",
        "sum_total":        "Total dû (£)",
        "sum_grand":        "TOTAL",
        "col_invoice_no":   "N° de facture",
        "col_company":      "Entreprise",
        "col_currency":     "Devise",
        "col_site":         "Site",
        "col_city":         "Ville",
        "col_postcode":     "Code postal",
        "col_meter_id":     "ID compteur",
        "col_period":       "Période de facturation",
        "col_period_start": "Début de période",
        "col_period_end":   "Fin de période",
        "col_issue_date":   "Date d'émission",
        "col_due_date":     "Date d'échéance",
        "col_prev_read":    "Relevé préc. (kWh)",
        "col_curr_read":    "Relevé actuel (kWh)",
        "col_consumption":  "Consommation (kWh)",
        "col_unit_price":   "Prix unit. (£/kWh)",
        "col_heat_cost":    "Coût thermique (£)",
        "col_capacity":     "Capacité (kW)",
        "col_cap_rate":     "Taux cap. (£/kW/mois)",
        "col_supplier_ef": "FE fournisseur (kg CO\u2082e/kWh)",
        "col_cap_charge":   "Frais de cap. (£)",
        "col_subtotal":     "Sous-total (£)",
        "col_vat":          "TVA (£)",
        "col_total":        "Total (£)",
    },
    "de": {
        "meta_period":      "Finanzzeitraum",
        "meta_start":       "Zeitraum Beginn",
        "meta_end":         "Zeitraum Ende",
        "meta_generated":   "Erstellt",
        "sum_company":      "Unternehmen",
        "sum_currency":     "Währung",
        "sum_sites":        "Standorte",
        "sum_invoices":     "Rechnungen",
        "sum_heat_cost":    "Wärmekosten (£)",
        "sum_cap_charge":   "Leistungsgebühr (£)",
        "sum_subtotal":     "Zwischensumme (£)",
        "sum_vat":          "MwSt. (£)",
        "sum_total":        "Gesamtbetrag (£)",
        "sum_grand":        "GESAMT",
        "col_invoice_no":   "Rechnungsnr.",
        "col_company":      "Unternehmen",
        "col_currency":     "Währung",
        "col_site":         "Standort",
        "col_city":         "Stadt",
        "col_postcode":     "Postleitzahl",
        "col_meter_id":     "Zähler-ID",
        "col_period":       "Abrechnungszeitraum",
        "col_period_start": "Zeitraum Beginn",
        "col_period_end":   "Zeitraum Ende",
        "col_issue_date":   "Ausstellungsdatum",
        "col_due_date":     "Fälligkeitsdatum",
        "col_prev_read":    "Vorh. Stand (kWh)",
        "col_curr_read":    "Akt. Stand (kWh)",
        "col_consumption":  "Verbrauch (kWh)",
        "col_unit_price":   "Einheitspreis (£/kWh)",
        "col_heat_cost":    "Wärmekosten (£)",
        "col_capacity":     "Leistung (kW)",
        "col_cap_rate":     "Leistungssatz (£/kW/Mo)",
        "col_supplier_ef": "EF Lieferant (kg CO\u2082e/kWh)",
        "col_cap_charge":   "Leistungsgebühr (£)",
        "col_subtotal":     "Zwischensumme (£)",
        "col_vat":          "MwSt. (£)",
        "col_total":        "Gesamt (£)",
    },
    "nl": {
        "meta_period":      "Financiële periode",
        "meta_start":       "Periode begin",
        "meta_end":         "Periode einde",
        "meta_generated":   "Gegenereerd",
        "sum_company":      "Bedrijf",
        "sum_currency":     "Valuta",
        "sum_sites":        "Locaties",
        "sum_invoices":     "Facturen",
        "sum_heat_cost":    "Warmtekosten (£)",
        "sum_cap_charge":   "Vermogenstoeslag (£)",
        "sum_subtotal":     "Subtotaal (£)",
        "sum_vat":          "BTW (£)",
        "sum_total":        "Totaal (£)",
        "sum_grand":        "TOTAAL",
        "col_invoice_no":   "Factuurnr.",
        "col_company":      "Bedrijf",
        "col_currency":     "Valuta",
        "col_site":         "Locatie",
        "col_city":         "Stad",
        "col_postcode":     "Postcode",
        "col_meter_id":     "Meter-ID",
        "col_period":       "Facturatieperiode",
        "col_period_start": "Periode begin",
        "col_period_end":   "Periode einde",
        "col_issue_date":   "Uitgiftedatum",
        "col_due_date":     "Vervaldatum",
        "col_prev_read":    "Vorige stand (kWh)",
        "col_curr_read":    "Huidige stand (kWh)",
        "col_consumption":  "Verbruik (kWh)",
        "col_unit_price":   "Eenheidsprijs (£/kWh)",
        "col_heat_cost":    "Warmtekosten (£)",
        "col_capacity":     "Vermogen (kW)",
        "col_cap_rate":     "Vermogenstarief (£/kW/mnd)",
        "col_supplier_ef": "EF leverancier (kg CO\u2082e/kWh)",
        "col_cap_charge":   "Vermogenstoeslag (£)",
        "col_subtotal":     "Subtotaal (£)",
        "col_vat":          "BTW (£)",
        "col_total":        "Totaal (£)",
    },
}

ELECTRICITY_TRANSLATIONS: dict[str, dict[str, str]] = {
    "en": {
        "xl_meta_period": "Financial Period",
        "xl_meta_start": "Period Start",
        "xl_meta_end": "Period End",
        "xl_meta_generated": "Generated",
        "xl_sum_company": "Company",
        "xl_sum_sites": "Sites",
        "xl_sum_qty": "Total Consumption",
        "xl_sum_cost": "Total Cost",
        "xl_sum_emissions_t": "Total tCO\u2082e",
        "xl_sum_grand": "TOTAL",
        "xl_col_ref": "Reference",
        "xl_col_company": "Company",
        "xl_col_currency": "Currency",
        "xl_col_site": "Site",
        "xl_col_period": "Billing Period",
        "xl_col_city": "City",
        "xl_col_postcode": "Postcode",
        "xl_col_meter_id": "Meter ID",
        "xl_col_supplier_ef": "Supplier EF (kg CO\u2082e/kWh)",
        "xl_col_unit": "Unit",
        "xl_col_start_read": "Start Reading",
        "xl_col_end_read": "End Reading",
        "xl_col_total_qty": "Total Quantity",
        "xl_col_total_cost": "Total Cost",
        "xl_col_emissions_kg": "Emissions (kg CO\u2082e)",
        "xl_col_emissions_t": "Emissions (tCO\u2082e)",
        "xl_tariff_name": "Tariff Name",
        "xl_tariff_qty": "Quantity",
        "xl_tariff_rate": "Unit Cost",
        "xl_tariff_cost": "Cost",
        "sm_col_consumption": "Consumption",
        "sm_col_tariff_type": "Tariff Type",
        "sm_col_timestamp": "Timestamp",
        "sm_col_import_kwh": "Import kWh",
        "sm_col_export_kwh": "Export kWh",
        "sm_col_end_reading": "End Reading",
    },
    "fr": {
        "xl_meta_period": "P\u00e9riode financi\u00e8re",
        "xl_meta_start": "D\u00e9but de p\u00e9riode",
        "xl_meta_end": "Fin de p\u00e9riode",
        "xl_meta_generated": "G\u00e9n\u00e9r\u00e9",
        "xl_sum_company": "Entreprise",
        "xl_sum_sites": "Sites",
        "xl_sum_qty": "Consommation totale",
        "xl_sum_cost": "Co\u00fbt total",
        "xl_sum_emissions_t": "Total tCO\u2082e",
        "xl_sum_grand": "TOTAL",
        "xl_col_ref": "R\u00e9f\u00e9rence",
        "xl_col_company": "Entreprise",
        "xl_col_currency": "Devise",
        "xl_col_site": "Site",
        "xl_col_period": "P\u00e9riode de facturation",
        "xl_col_city": "Ville",
        "xl_col_postcode": "Code postal",
        "xl_col_meter_id": "ID compteur",
        "xl_col_supplier_ef": "FE fournisseur (kg CO\u2082e/kWh)",
        "xl_col_unit": "Unit\u00e9",
        "xl_col_start_read": "Relev\u00e9 initial",
        "xl_col_end_read": "Relev\u00e9 final",
        "xl_col_total_qty": "Quantit\u00e9 totale",
        "xl_col_total_cost": "Co\u00fbt total",
        "xl_col_emissions_kg": "\u00c9missions (kg CO\u2082e)",
        "xl_col_emissions_t": "\u00c9missions (tCO\u2082e)",
        "xl_tariff_name": "Nom du tarif",
        "xl_tariff_qty": "Quantit\u00e9",
        "xl_tariff_rate": "Co\u00fbt unitaire",
        "xl_tariff_cost": "Co\u00fbt",
        "sm_col_consumption": "Consommation",
        "sm_col_tariff_type": "Type de tarif",
        "sm_col_timestamp": "Horodatage",
        "sm_col_import_kwh": "kWh import\u00e9s",
        "sm_col_export_kwh": "kWh export\u00e9s",
        "sm_col_end_reading": "Relev\u00e9 final",
    },
    "de": {
        "xl_meta_period": "Finanzzeitraum",
        "xl_meta_start": "Zeitraum Beginn",
        "xl_meta_end": "Zeitraum Ende",
        "xl_meta_generated": "Erstellt",
        "xl_sum_company": "Unternehmen",
        "xl_sum_sites": "Standorte",
        "xl_sum_qty": "Gesamtverbrauch",
        "xl_sum_cost": "Gesamtkosten",
        "xl_sum_emissions_t": "Gesamt tCO\u2082e",
        "xl_sum_grand": "GESAMT",
        "xl_col_ref": "Referenz",
        "xl_col_company": "Unternehmen",
        "xl_col_currency": "Währung",
        "xl_col_site": "Standort",
        "xl_col_period": "Abrechnungszeitraum",
        "xl_col_city": "Stadt",
        "xl_col_postcode": "Postleitzahl",
        "xl_col_meter_id": "Z\u00e4hler-ID",
        "xl_col_supplier_ef": "EF Lieferant (kg CO\u2082e/kWh)",
        "xl_col_unit": "Einheit",
        "xl_col_start_read": "Anfangsz\u00e4hlerstand",
        "xl_col_end_read": "Endz\u00e4hlerstand",
        "xl_col_total_qty": "Gesamtmenge",
        "xl_col_total_cost": "Gesamtkosten",
        "xl_col_emissions_kg": "Emissionen (kg CO\u2082e)",
        "xl_col_emissions_t": "Emissionen (tCO\u2082e)",
        "xl_tariff_name": "Tarifname",
        "xl_tariff_qty": "Menge",
        "xl_tariff_rate": "Einheitspreis",
        "xl_tariff_cost": "Kosten",
        "sm_col_consumption": "Verbrauch",
        "sm_col_tariff_type": "Tariftyp",
        "sm_col_timestamp": "Zeitstempel",
        "sm_col_import_kwh": "Import kWh",
        "sm_col_export_kwh": "Export kWh",
        "sm_col_end_reading": "Endstand",
    },
    "nl": {
        "xl_meta_period": "Financi\u00eble periode",
        "xl_meta_start": "Periode begin",
        "xl_meta_end": "Periode einde",
        "xl_meta_generated": "Gegenereerd",
        "xl_sum_company": "Bedrijf",
        "xl_sum_sites": "Locaties",
        "xl_sum_qty": "Totaal verbruik",
        "xl_sum_cost": "Totale kosten",
        "xl_sum_emissions_t": "Totaal tCO\u2082e",
        "xl_sum_grand": "TOTAAL",
        "xl_col_ref": "Referentie",
        "xl_col_company": "Bedrijf",
        "xl_col_currency": "Valuta",
        "xl_col_site": "Locatie",
        "xl_col_period": "Facturatieperiode",
        "xl_col_city": "Stad",
        "xl_col_postcode": "Postcode",
        "xl_col_meter_id": "Meter-ID",
        "xl_col_supplier_ef": "EF leverancier (kg CO\u2082e/kWh)",
        "xl_col_unit": "Eenheid",
        "xl_col_start_read": "Beginmeterstand",
        "xl_col_end_read": "Eindmeterstand",
        "xl_col_total_qty": "Totale hoeveelheid",
        "xl_col_total_cost": "Totale kosten",
        "xl_col_emissions_kg": "Emissies (kg CO\u2082e)",
        "xl_col_emissions_t": "Emissies (tCO\u2082e)",
        "xl_tariff_name": "Tariefnaam",
        "xl_tariff_qty": "Hoeveelheid",
        "xl_tariff_rate": "Eenheidsprijs",
        "xl_tariff_cost": "Kosten",
        "sm_col_consumption": "Verbruik",
        "sm_col_tariff_type": "Tarieftype",
        "sm_col_timestamp": "Tijdstempel",
        "sm_col_import_kwh": "Import kWh",
        "sm_col_export_kwh": "Export kWh",
        "sm_col_end_reading": "Eindstand",
    },
}


# ── palette ───────────────────────────────────────────────────────────────────
_WHITE = "FFFFFF"
_DARK = "1F2328"
_MID = "5A6066"
_ALT = "F7F9FC"
_SOFT = "DCEBF5"
_BORDER = "C9CDD2"

# ── style primitives ──────────────────────────────────────────────────────────

def _font(bold: bool = False, color: str = _DARK, size: int = 10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _border() -> Border:
    side = Side(style="thin", color=_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _align(horizontal: str = "left") -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center")


def _header_cell(cell, text: str, accent: str) -> None:
    cell.value = text
    cell.font = _font(bold=True, color=_WHITE)
    cell.fill = _fill(accent)
    cell.border = _border()
    cell.alignment = _align("center")


def _data_cell(cell, value: Any, fmt: str | None = None, alt: bool = False, bold: bool = False) -> None:
    cell.value = value
    cell.font = _font(bold=bold)
    cell.fill = _fill(_ALT if alt else _WHITE)
    cell.border = _border()
    cell.alignment = _align("right" if isinstance(value, (int, float, Decimal)) else "left")
    if fmt:
        cell.number_format = fmt


_HEAT_DETAIL_COLUMN_SPECS: list[tuple[str, str, int, str | None, str | None, Any]] = [
    ("invoice_no", "col_invoice_no", 20, None, "invoice_no", lambda co, si, rec: rec["invoice_no"]),
    ("company", "col_company", 30, None, None, lambda co, si, rec: co["label"]),
    ("site", "col_site", 20, None, "site_label", lambda co, si, rec: si["label"]),
    ("city", "col_city", 14, None, "city", lambda co, si, rec: rec["city"]),
    ("postcode", "col_postcode", 11, None, "postcode", lambda co, si, rec: rec["postcode"]),
    ("meter_id", "col_meter_id", 22, None, "meter_id", lambda co, si, rec: rec["meter_id"]),
    ("period", "col_period", 18, None, None, lambda co, si, rec: rec["billing_period_label"]),
    ("period_start", "col_period_start", 14, "DD MMM YYYY", None, lambda co, si, rec: rec["period_start"]),
    ("period_end", "col_period_end", 14, "DD MMM YYYY", None, lambda co, si, rec: rec["period_end"]),
    ("issue_date", "col_issue_date", 14, "DD MMM YYYY", None, lambda co, si, rec: rec["issue_date"]),
    ("due_date", "col_due_date", 14, "DD MMM YYYY", None, lambda co, si, rec: rec["due_date"]),
    ("prev_read", "col_prev_read", 20, "#,##0", "prev_read", lambda co, si, rec: rec["prev_read"]),
    ("curr_read", "col_curr_read", 20, "#,##0", "curr_read", lambda co, si, rec: rec["curr_read"]),
    ("consumption", "col_consumption", 20, "#,##0", "consumption", lambda co, si, rec: rec["consumption"]),
    ("unit_price", "col_unit_price", 20, "0.000", "unit_price", lambda co, si, rec: rec["unit_price"] if isinstance(rec["unit_price"], str) else float(rec["unit_price"])),
    ("heat_cost", "col_heat_cost", 16, "#,##0.00", "heat_cost", lambda co, si, rec: rec["heat_cost"] if isinstance(rec["heat_cost"], str) else float(rec["heat_cost"])),
    ("capacity", "col_capacity", 14, "#,##0", "capacity_kw", lambda co, si, rec: rec["capacity_kw"]),
    ("cap_rate", "col_cap_rate", 20, "0.00", "capacity_rate", lambda co, si, rec: rec["capacity_rate"] if isinstance(rec["capacity_rate"], str) else float(rec["capacity_rate"])),
    ("supplier_ef", "col_supplier_ef", 24, "0.0000", "supplier_ef", lambda co, si, rec: rec["supplier_ef"] if isinstance(rec["supplier_ef"], str) else float(rec["supplier_ef"])),
    ("cap_charge", "col_cap_charge", 20, "#,##0.00", "capacity_charge", lambda co, si, rec: rec["capacity_charge"] if isinstance(rec["capacity_charge"], str) else float(rec["capacity_charge"])),
    ("subtotal", "col_subtotal", 16, "#,##0.00", None, lambda co, si, rec: rec["subtotal"] if isinstance(rec["subtotal"], str) else float(rec["subtotal"])),
    ("vat", "col_vat", 13, "#,##0.00", None, lambda co, si, rec: rec["vat"] if isinstance(rec["vat"], str) else float(rec["vat"])),
    ("total", "col_total", 16, "#,##0.00", None, lambda co, si, rec: rec["total"] if isinstance(rec["total"], str) else float(rec["total"])),
    ("currency", "col_currency", 10, None, None, lambda co, si, rec: currency_code(co.get("currency"))),
]

_HEAT_DETAIL_COLUMN_MAP = {
    field_id: {
        "header_key": header_key,
        "width": width,
        "fmt": fmt,
        "blank_field": blank_field,
        "accessor": accessor,
    }
    for field_id, header_key, width, fmt, blank_field, accessor in _HEAT_DETAIL_COLUMN_SPECS
}

_ELECTRICITY_CORE_COLUMN_SPECS: list[tuple[str, str, int, str | None, Any]] = [
    ("reference", "xl_col_ref", 30, None, lambda co, si: si["ref_no"]),
    ("company", "xl_col_company", 28, None, lambda co, si: co["label"]),
    ("site", "xl_col_site", 22, None, lambda co, si: si["label"]),
    ("period", "xl_col_period", 18, None, lambda co, si: si["billing_period_label"]),
    ("city", "xl_col_city", 16, None, lambda co, si: si["city"]),
    ("postcode", "xl_col_postcode", 10, None, lambda co, si: si["postcode"]),
    ("meter_id", "xl_col_meter_id", 24, None, lambda co, si: si["meter_id"]),
    ("supplier_ef", "xl_col_supplier_ef", 16, "#,##0.0000", lambda co, si: si["supplier_ef"] if isinstance(si["supplier_ef"], str) else float(si["supplier_ef"])),
    ("unit", "xl_col_unit", 8, None, lambda co, si: si["unit"]),
    ("start_read", "xl_col_start_read", 14, "#,##0", lambda co, si: si["start_reading"]),
    ("end_read", "xl_col_end_read", 14, "#,##0", lambda co, si: si["end_reading"]),
    ("total_qty", "xl_col_total_qty", 14, "#,##0.00", lambda co, si: si["total_quantity"] if isinstance(si["total_quantity"], str) else float(si["total_quantity"])),
    ("total_cost", "xl_col_total_cost", 14, "#,##0.00", lambda co, si: si["total_cost"] if isinstance(si["total_cost"], str) else float(si["total_cost"])),
    ("currency", "xl_col_currency", 10, None, lambda co, si: currency_code(co.get("currency"))),
    ("emissions_kg", "xl_col_emissions_kg", 16, "#,##0.00", lambda co, si: si["emissions_kg"] if isinstance(si.get("emissions_kg"), str) else float(si["emissions_kg"])),
    ("emissions_t", "xl_col_emissions_t", 16, "#,##0.00", lambda co, si: si["emissions_t"] if isinstance(si.get("emissions_t"), str) else float(si["emissions_t"])),
]

_ELECTRICITY_CORE_COLUMN_MAP = {
    field_id: {"header_key": header_key, "width": width, "fmt": fmt, "accessor": accessor}
    for field_id, header_key, width, fmt, accessor in _ELECTRICITY_CORE_COLUMN_SPECS
}

_SMART_METER_MONTHLY_SPECS: list[tuple[str, str, int, str | None, Any]] = [
    ("meter_id", "xl_col_meter_id", 22, None, lambda row: row["meter_id"]),
    ("site", "xl_col_site", 24, None, lambda row: row["site_label"]),
    ("period", "xl_col_period", 18, None, lambda row: row["period_label"]),
    ("start_read", "xl_col_start_read", 14, "#,##0", lambda row: row["start_reading"]),
    ("end_read", "xl_col_end_read", 14, "#,##0", lambda row: row["end_reading"]),
    ("consumption", "sm_col_consumption", 14, "#,##0.00", lambda row: row["consumption"]),
    ("unit", "xl_col_unit", 10, None, lambda row: row["unit"]),
    ("tariff_type", "sm_col_tariff_type", 18, None, lambda row: row["tariff_type"]),
    ("tariff_cost", "xl_tariff_cost", 12, "#,##0.00", lambda row: row["cost"]),
    ("currency", "xl_col_currency", 10, None, lambda row: row["currency"]),
]

_SMART_METER_INTERVAL_SPECS: dict[str, list[tuple[str, str, int, str | None, Any]]] = {
    "consumption_diff": [
        ("meter_id", "xl_col_meter_id", 22, None, lambda row: row["meter_id"]),
        ("timestamp", "sm_col_timestamp", 24, None, lambda row: row["timestamp"]),
        ("import_kwh", "sm_col_import_kwh", 14, "0.0000", lambda row: row["import_kwh"]),
        ("export_kwh", "sm_col_export_kwh", 14, "0.0000", lambda row: row["export_kwh"]),
        ("unit", "xl_col_unit", 10, None, lambda row: row["unit"]),
    ],
    "cumulative_end_reading": [
        ("meter_id", "xl_col_meter_id", 22, None, lambda row: row["meter_id"]),
        ("timestamp", "sm_col_timestamp", 24, None, lambda row: row["timestamp"]),
        ("end_read", "sm_col_end_reading", 16, "0.0000", lambda row: row["end_reading"]),
        ("unit", "xl_col_unit", 10, None, lambda row: row["unit"]),
    ],
}


# ── helpers ───────────────────────────────────────────────────────────────────

def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Return a valid Excel sheet name (max 31 chars, no forbidden characters)."""
    sanitised = re.sub(r'[\\/?*\[\]:]', '', name).strip()
    return (sanitised or "Company")[:max_len]


def _sections_by_company(sections: list[dict]) -> dict[str, list[dict]]:
    """Group sections by company label, preserving insertion order."""
    grouped: dict[str, list[dict]] = {}
    for section in sections:
        label = section["company"]["label"]
        grouped.setdefault(label, []).append(section)
    return grouped


def _currency_label_for_sections(sections: list[dict]) -> str:
    codes = {currency_code(section["company"].get("currency")) for section in sections}
    return next(iter(codes)) if len(codes) == 1 else "Currency"


def _replace_currency_labels(strings: dict[str, str], currency_label: str) -> dict[str, str]:
    return {key: value.replace("£", currency_label) for key, value in strings.items()}


def _layout_plan(config: dict) -> dict:
    return config.get("document", {}).get("layout_plan", {})


def _distractor_plan(config: dict):
    return config.get("document", {}).get("distractor_plan")


def _ordered_ids(plan: dict, available_ids: list[str]) -> list[str]:
    excluded = set(plan.get("excluded_fields") or [])
    available_ids = [field_id for field_id in available_ids if field_id not in excluded]
    requested = list(plan.get("column_order") or [])
    if not requested:
        return available_ids
    ordered = [field_id for field_id in requested if field_id in available_ids]
    ordered.extend(field_id for field_id in available_ids if field_id not in ordered)
    return ordered


def _header_text(strings: dict[str, str], plan: dict, field_id: str, header_key: str) -> str:
    return str((plan.get("header_aliases") or {}).get(field_id, strings[header_key]))


def _write_sheet_preamble(ws, plan: dict) -> int:
    row = 1
    for preamble_row in plan.get("preamble_rows") or []:
        for col_idx, value in enumerate(preamble_row, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1
    row += int(plan.get("header_row_offset", 0))
    return row


def _augment_columns_with_distractors(columns: list[dict[str, Any]], distractor_plan) -> list[dict[str, Any]]:
    if not distractor_plan or not getattr(distractor_plan, "enabled", False):
        return columns

    augmented = list(columns)
    for field in distractor_plan.tabular_fields:
        insert_at = len(augmented)
        anchors = [column.get("field_id") for column in augmented]
        if field.anchor in anchors:
            anchor_index = anchors.index(field.anchor)
            insert_at = anchor_index + 1 if field.position == "after" else anchor_index
        augmented.insert(
            insert_at,
            {
                "kind": "distractor",
                "field_id": field.field_id,
                "label": field.label,
                "width": 16,
                "fmt": None,
                "distractor_field": field,
            },
        )
    return augmented


def _heat_row_key(company: dict, site: dict, rec: dict) -> str:
    del company, site
    return str(rec.get("invoice_no") or rec.get("billing_period_label") or "heat-row")


def _electricity_row_key(company: dict, site: dict) -> str:
    del company
    return str(site.get("ref_no") or site.get("meter_id") or site.get("billing_period_label") or "electricity-row")


def _smart_meter_row_key(row: dict) -> str:
    return str(row.get("timestamp") or row.get("period_label") or row.get("meter_id") or "smart-meter-row")


def _heat_detail_columns(strings: dict, plan: dict, distractor_plan) -> list[dict[str, Any]]:
    ordered_field_ids = _ordered_ids(plan, list(_HEAT_DETAIL_COLUMN_MAP))
    columns = [
        {
            "kind": "base",
            "field_id": field_id,
            "label": _header_text(strings, plan, field_id, _HEAT_DETAIL_COLUMN_MAP[field_id]["header_key"]),
            **_HEAT_DETAIL_COLUMN_MAP[field_id],
        }
        for field_id in ordered_field_ids
    ]
    return _augment_columns_with_distractors(columns, distractor_plan)


def _electricity_detail_columns(strings: dict, plan: dict, max_tariffs: int, distractor_plan) -> list[dict[str, Any]]:
    ordered_field_ids = _ordered_ids(plan, list(_ELECTRICITY_CORE_COLUMN_MAP))
    columns: list[dict[str, Any]] = []
    tariff_inserted = False
    for field_id in ordered_field_ids:
        spec = _ELECTRICITY_CORE_COLUMN_MAP[field_id]
        columns.append({
            "kind": "core",
            "field_id": field_id,
            "label": _header_text(strings, plan, field_id, spec["header_key"]),
            **spec,
        })
        if field_id == "total_qty" and plan.get("tariff_block_position") == "after_total_qty":
            columns.extend(_electricity_tariff_columns(strings, max_tariffs))
            tariff_inserted = True
    if not tariff_inserted:
        columns.extend(_electricity_tariff_columns(strings, max_tariffs))
    return _augment_columns_with_distractors(columns, distractor_plan)


def _electricity_tariff_columns(strings: dict, max_tariffs: int) -> list[dict[str, Any]]:
    columns: list[dict[str, Any]] = []
    for idx in range(max_tariffs):
        prefix = f"Tariff {idx + 1}"
        columns.extend([
            {"kind": "tariff", "idx": idx, "part": "name", "label": f"{prefix}: {strings['xl_tariff_name']}", "width": 26, "fmt": None},
            {"kind": "tariff", "idx": idx, "part": "qty", "label": f"{prefix}: {strings['xl_tariff_qty']}", "width": 12, "fmt": "#,##0.00"},
            {"kind": "tariff", "idx": idx, "part": "rate", "label": f"{prefix}: {strings['xl_tariff_rate']}", "width": 12, "fmt": "#,##0.0000"},
            {"kind": "tariff", "idx": idx, "part": "cost", "label": f"{prefix}: {strings['xl_tariff_cost']}", "width": 12, "fmt": "#,##0.00"},
        ])
    return columns


def _smart_meter_columns(strings: dict, plan: dict, mode: str, value_mode: str, distractor_plan) -> list[dict[str, Any]]:
    base_specs = _SMART_METER_MONTHLY_SPECS if mode != "interval" else _SMART_METER_INTERVAL_SPECS.get(value_mode, _SMART_METER_INTERVAL_SPECS["consumption_diff"])
    spec_map = {
        field_id: {"header_key": header_key, "width": width, "fmt": fmt, "accessor": accessor}
        for field_id, header_key, width, fmt, accessor in base_specs
    }
    ordered_field_ids = _ordered_ids(plan, list(spec_map))
    columns = [
        {
            "kind": "base",
            "field_id": field_id,
            "label": _header_text(strings, plan, field_id, spec_map[field_id]["header_key"]),
            **spec_map[field_id],
        }
        for field_id in ordered_field_ids
    ]
    return _augment_columns_with_distractors(columns, distractor_plan)


def _reorder_workbook_sheets(workbook, desired_titles: list[str]) -> None:
    if not desired_titles:
        return
    title_map = {sheet.title: sheet for sheet in workbook.worksheets}
    ordered = [title_map[title] for title in desired_titles if title in title_map]
    ordered.extend(sheet for sheet in workbook.worksheets if sheet not in ordered)
    workbook._sheets = ordered


# ── public API ────────────────────────────────────────────────────────────────

def _generate_heat_xlsx(
    config: dict,
    sections: list[dict],
    blank_fields: set[str] | None = None,
    split_by_company: bool = False,
    include_summary: bool = True,
) -> bytes:
    """Build a styled XLSX workbook from billing sections and return bytes.

    blank_fields: record field names whose cells should be left empty (QA testing).
    split_by_company: when True, create one detail sheet per company instead of
                      a single combined "Billing Detail" sheet.
    """
    lang = config["document"].get("language", "en")
    strings = _replace_currency_labels(TRANSLATIONS.get(lang, TRANSLATIONS["en"]), _currency_label_for_sections(sections))
    plan = _layout_plan(config)
    distractor_plan = _distractor_plan(config)

    default_accent = sections[0]["company"]["accent"].lstrip("#") if sections else "1E5B88"
    omit = blank_fields or set()

    wb = openpyxl.Workbook()

    if include_summary:
        wb.active.title = "Summary"
        _build_summary_sheet(wb.active, config, sections, default_accent, strings)
        detail_seed_sheet = None
    else:
        detail_seed_sheet = wb.active

    if split_by_company:
        grouped = _sections_by_company(sections)
        ordered_labels = list(plan.get("company_sheet_order") or grouped.keys())
        by_company = [(label, grouped[label]) for label in ordered_labels if label in grouped]
        for idx, (label, co_sections) in enumerate(by_company):
            accent = co_sections[0]["company"]["accent"].lstrip("#")
            sheet_name = _safe_sheet_name(label)
            target_sheet = detail_seed_sheet if idx == 0 and detail_seed_sheet is not None else wb.create_sheet(sheet_name)
            target_sheet.title = sheet_name
            _build_detail_sheet(target_sheet, co_sections, accent, omit, strings, plan, distractor_plan)
    else:
        target_sheet = detail_seed_sheet if detail_seed_sheet is not None else wb.create_sheet("Billing Detail")
        target_sheet.title = "Billing Detail"
        _build_detail_sheet(target_sheet, sections, default_accent, omit, strings, plan, distractor_plan)

    if include_summary and plan.get("sheet_order"):
        desired_titles = []
        for sheet_id in plan["sheet_order"]:
            if sheet_id == "summary":
                desired_titles.append("Summary")
            elif sheet_id == "detail":
                desired_titles.append("Billing Detail")
            else:
                desired_titles.append(_safe_sheet_name(sheet_id))
        _reorder_workbook_sheets(wb, desired_titles)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_electricity_xlsx(config: dict, sections: list[dict], include_summary: bool = True) -> bytes:
    lang = config["document"].get("language", "en")
    strings = ELECTRICITY_TRANSLATIONS.get(lang, ELECTRICITY_TRANSLATIONS["en"])
    financial_period = config["financial_period"]
    plan = _layout_plan(config)
    distractor_plan = _distractor_plan(config)

    accent_hex = sections[0]["company"]["accent"] if sections else "#1E5B88"
    accent_r, accent_g, accent_b = (int(accent_hex[i:i + 2], 16) for i in (1, 3, 5))

    def header_fill(hex_color: str) -> PatternFill:
        red, green, blue = (int(hex_color[i:i + 2], 16) for i in (1, 3, 5))
        return PatternFill("solid", fgColor=f"{red:02X}{green:02X}{blue:02X}")

    def header_font(white_text: bool = True) -> Font:
        return Font(name="Calibri", bold=True, color="FFFFFF" if white_text else "1F2328", size=9)

    def thin_border() -> Border:
        side = Side(style="thin", color="D5DADF")
        return Border(left=side, right=side, top=side, bottom=side)

    workbook = openpyxl.Workbook()
    hdr_fill = header_fill(accent_hex)
    hdr_font = header_font()
    border = thin_border()

    if include_summary:
        summary = workbook.active
        summary.title = "Summary"

        summary["A1"] = strings["xl_meta_period"]
        summary["B1"] = financial_period["label"]
        summary["A2"] = strings["xl_meta_start"]
        summary["B2"] = financial_period["start_date"].strftime("%d %b %Y") if hasattr(financial_period["start_date"], "strftime") else str(financial_period["start_date"])
        summary["A3"] = strings["xl_meta_end"]
        summary["B3"] = financial_period["end_date"].strftime("%d %b %Y") if hasattr(financial_period["end_date"], "strftime") else str(financial_period["end_date"])
        summary["A4"] = strings["xl_meta_generated"]
        summary["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        for row in range(1, 5):
            summary.cell(row, 1).font = Font(name="Calibri", bold=True, size=9)

        summary_headers = [
            strings["xl_sum_company"],
            strings["xl_sum_sites"],
            strings["xl_sum_qty"],
            strings["xl_sum_cost"],
            strings["xl_col_currency"],
            strings["xl_sum_emissions_t"],
        ]
        header_row = 6
        for col_idx, header in enumerate(summary_headers, start=1):
            cell = summary.cell(header_row, col_idx, header)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        from collections import defaultdict

        by_company: dict[str, list] = defaultdict(list)
        for section in sections:
            by_company[section["company"]["label"]].append(section)

        grand_qty = Decimal("0")
        grand_cost = Decimal("0")
        grand_emissions = Decimal("0")

        data_row = header_row + 1
        for company_label, company_sections in by_company.items():
            site_count = len({section["site"]["_site_uid"] for section in company_sections})
            total_qty = sum(
                v for s in company_sections
                if not isinstance((v := s["site"]["total_quantity"]), str)
            )
            total_cost = sum(
                v for s in company_sections
                if not isinstance((v := s["site"]["total_cost"]), str)
            )
            total_emissions = sum(
                v for s in company_sections
                if not isinstance((v := s["site"]["emissions_t"]), str)
            )

            row_values = [
                company_label,
                site_count,
                float(total_qty),
                float(total_cost),
                currency_code(company_sections[0]["company"].get("currency")),
                float(total_emissions),
            ]
            for col_idx, value in enumerate(row_values, start=1):
                cell = summary.cell(data_row, col_idx, value)
                cell.border = border
                cell.font = Font(name="Calibri", size=9)
                if col_idx in {2, 3}:
                    cell.number_format = "#,##0"
                elif col_idx in {4, 6}:
                    cell.number_format = "#,##0.00"
            data_row += 1
            grand_qty += total_qty
            grand_cost += total_cost
            grand_emissions += total_emissions

        grand_fill = PatternFill("solid", fgColor=f"{accent_r:02X}{accent_g:02X}{accent_b:02X}")
        grand_values = [
            strings["xl_sum_grand"],
            len({section["site"]["_site_uid"] for section in sections}),
            float(grand_qty),
            float(grand_cost),
            "",
            float(grand_emissions),
        ]
        for col_idx, value in enumerate(grand_values, start=1):
            cell = summary.cell(data_row, col_idx, value)
            cell.fill = grand_fill
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
            cell.border = border
            if col_idx in {2, 3}:
                cell.number_format = "#,##0"
            elif col_idx in {4, 6}:
                cell.number_format = "#,##0.00"

        for col_idx, width in enumerate([30, 8, 18, 18, 10, 14], start=1):
            summary.column_dimensions[get_column_letter(col_idx)].width = width

        detail = workbook.create_sheet("Detail")
    else:
        detail = workbook.active
        detail.title = "Detail"
    max_tariffs = max((len(section["site"].get("tariffs", [])) for section in sections), default=0)

    detail_columns = _electricity_detail_columns(strings, plan, max_tariffs, distractor_plan)
    header_row = _write_sheet_preamble(detail, plan)
    for col_idx, column in enumerate(detail_columns, start=1):
        cell = detail.cell(header_row, col_idx, column["label"])
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        detail.column_dimensions[get_column_letter(col_idx)].width = column["width"]
    detail.freeze_panes = f"A{header_row + 1}"

    for row_idx, section in enumerate(sections, start=header_row + 1):
        company = section["company"]
        site = section["site"]
        tariffs = site.get("tariffs", [])

        for col_idx, column in enumerate(detail_columns, start=1):
            if column["kind"] == "core":
                value = column["accessor"](company, site)
            elif column["kind"] == "distractor":
                value = resolve_tabular_value(distractor_plan, column["distractor_field"], _electricity_row_key(company, site))
            else:
                tariff = tariffs[column["idx"]] if column["idx"] < len(tariffs) else None
                if tariff is None:
                    value = ""
                elif column["part"] == "name":
                    value = tariff["name"]
                elif column["part"] == "qty":
                    value = tariff["quantity"] if isinstance(tariff["quantity"], str) else float(tariff["quantity"])
                elif column["part"] == "rate":
                    value = tariff["unit_cost"] if isinstance(tariff["unit_cost"], str) else float(tariff["unit_cost"])
                else:
                    value = tariff["cost"] if isinstance(tariff["cost"], str) else float(tariff["cost"])
            cell = detail.cell(row_idx, col_idx, value)
            cell.border = border
            cell.font = Font(name="Calibri", size=9)
            if column.get("fmt") and isinstance(value, (int, float, Decimal)):
                cell.number_format = column["fmt"]

    if include_summary and plan.get("sheet_order"):
        desired_titles = ["Summary" if sheet_id == "summary" else "Detail" for sheet_id in plan["sheet_order"]]
        _reorder_workbook_sheets(workbook, desired_titles)

    buf = BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _generate_smart_meter_xlsx(config: dict, sections: list[dict]) -> bytes:
    from generators.electricity_generator import build_smart_meter_rows

    lang = config["document"].get("language", "en")
    strings = ELECTRICITY_TRANSLATIONS.get(lang, ELECTRICITY_TRANSLATIONS["en"])
    mode = str(config["document"].get("smart_meter_data_granularity", "monthly")).lower()
    value_mode = str(config["document"].get("smart_meter_interval_value_mode", "consumption_diff")).lower()
    plan = _layout_plan(config)
    distractor_plan = _distractor_plan(config)
    rows = build_smart_meter_rows(config, sections)

    accent = config["companies"][0].get("accent", "#1E5B88") if config.get("companies") else "#1E5B88"
    accent_fill = accent.lstrip("#").upper()
    if len(accent_fill) == 6:
        accent_fill = f"FF{accent_fill}"

    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Smart Meter Data"
    columns = _smart_meter_columns(strings, plan, mode, value_mode, distractor_plan)
    header_row = _write_sheet_preamble(ws, plan)
    ws.freeze_panes = f"A{header_row + 1}"

    for col_idx, column in enumerate(columns, start=1):
        _header_cell(ws.cell(row=header_row, column=col_idx), column["label"], accent_fill)
        ws.column_dimensions[get_column_letter(col_idx)].width = column["width"]

    for row_idx, row in enumerate(rows, start=header_row + 1):
        alt = row_idx % 2 == 0
        for col_idx, column in enumerate(columns, start=1):
            if column["kind"] == "distractor":
                value = resolve_tabular_value(distractor_plan, column["distractor_field"], _smart_meter_row_key(row))
            else:
                value = column["accessor"](row)
            _data_cell(ws.cell(row=row_idx, column=col_idx), value, fmt=column["fmt"], alt=alt)

    buf = BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _generate_heat_supplier_portal_xlsx(
    config: dict,
    sections: list[dict],
    blank_fields: set[str] | None = None,
    split_by_company: bool = False,
    include_summary: bool = True,
) -> bytes:
    return _generate_heat_xlsx(
        config,
        sections,
        blank_fields=blank_fields,
        split_by_company=split_by_company,
        include_summary=include_summary,
    )


def _generate_electricity_supplier_portal_xlsx(
    config: dict,
    sections: list[dict],
    include_summary: bool = True,
) -> bytes:
    return _generate_electricity_xlsx(config, sections, include_summary=include_summary)


# ── summary sheet ─────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, config: dict, sections: list[dict], accent: str, strings: dict) -> None:
    # Title banner
    ws.merge_cells("A1:I1")
    ws["A1"].value = config["document"]["title"]
    ws["A1"].font = _font(bold=True, color=_WHITE, size=14)
    ws["A1"].fill = _fill(accent)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 30

    # Metadata
    meta = [
        (strings["meta_period"],    config["financial_period"]["label"]),
        (strings["meta_start"],     config["financial_period"]["start_date"].strftime("%d %b %Y")),
        (strings["meta_end"],       config["financial_period"]["end_date"].strftime("%d %b %Y")),
        (strings["meta_generated"], datetime.now().strftime("%d %b %Y %H:%M")),
    ]
    for offset, (label, value) in enumerate(meta):
        row = 2 + offset
        ws.cell(row=row, column=1, value=label).font = _font(bold=True, color=_MID)
        ws.cell(row=row, column=2, value=value).font = _font()

    # Company summary table
    tbl_row = 2 + len(meta) + 1
    summary_headers = [
        strings["sum_company"], strings["sum_sites"], strings["sum_invoices"],
        strings["sum_heat_cost"], strings["sum_cap_charge"],
        strings["sum_subtotal"], strings["sum_vat"],     strings["sum_total"], strings["sum_currency"],
    ]
    for col, h in enumerate(summary_headers, start=1):
        _header_cell(ws.cell(row=tbl_row, column=col), h, accent)

    # Aggregate per company
    totals: dict[str, dict] = {}
    for section in sections:
        key = section["company"]["label"]
        if key not in totals:
            totals[key] = {
                "sites": set(),
                "currency": currency_code(section["company"].get("currency")),
                "invoices": 0,
                "heat_cost": Decimal("0"),
                "capacity_charge": Decimal("0"),
                "subtotal": Decimal("0"),
                "vat": Decimal("0"),
                "total": Decimal("0"),
            }
        t = totals[key]
        t["sites"].add(section["site"]["label"])
        for rec in section["records"]:
            t["invoices"] += 1
            for _k in ("heat_cost", "capacity_charge", "subtotal", "vat", "total"):
                if not isinstance(rec[_k], str):
                    t[_k] += rec[_k]

    money_fmt = "#,##0.00"
    for i, (company, t) in enumerate(totals.items()):
        row = tbl_row + 1 + i
        alt = i % 2 == 1
        row_values = [
            (company,               None),
            (len(t["sites"]),       "#,##0"),
            (t["invoices"],         "#,##0"),
            (float(t["heat_cost"]), money_fmt),
            (float(t["capacity_charge"]), money_fmt),
            (float(t["subtotal"]),  money_fmt),
            (float(t["vat"]),       money_fmt),
            (float(t["total"]),     money_fmt),
            (t["currency"],         None),
        ]
        for col, (val, fmt) in enumerate(row_values, start=1):
            _data_cell(ws.cell(row=row, column=col), val, fmt=fmt, alt=alt)

    # Grand total row
    grand_row = tbl_row + 1 + len(totals)
    grand_values = [
        (strings["sum_grand"], None),
        ("",      None),
        (sum(t["invoices"] for t in totals.values()), "#,##0"),
        (float(sum(t["heat_cost"]        for t in totals.values())), money_fmt),
        (float(sum(t["capacity_charge"]  for t in totals.values())), money_fmt),
        (float(sum(t["subtotal"]         for t in totals.values())), money_fmt),
        (float(sum(t["vat"]              for t in totals.values())), money_fmt),
        (float(sum(t["total"]            for t in totals.values())), money_fmt),
        ("",      None),
    ]
    for col, (val, fmt) in enumerate(grand_values, start=1):
        c = ws.cell(row=grand_row, column=col, value=val)
        c.font = _font(bold=True)
        c.fill = _fill(_SOFT)
        c.border = _border()
        c.alignment = _align("right" if col > 2 else "left")
        if fmt:
            c.number_format = fmt

    # Column widths
    for col, width in enumerate([30, 8, 10, 20, 22, 18, 14, 20, 10], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"


# ── detail sheet ──────────────────────────────────────────────────────────────

def _build_detail_sheet(ws, sections: list[dict], accent: str, blank_fields: set[str], strings: dict, plan: dict, distractor_plan) -> None:
    # Headers
    columns = _heat_detail_columns(strings, plan, distractor_plan)
    header_row = _write_sheet_preamble(ws, plan)
    for col, column in enumerate(columns, start=1):
        _header_cell(ws.cell(row=header_row, column=col), column["label"], accent)
        ws.column_dimensions[get_column_letter(col)].width = column["width"]
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = f"A{header_row + 1}"

    # Each entry: (value, record_field_name_or_None, fmt)
    # record_field_name is checked against blank_fields to omit the cell value.
    row = header_row + 1
    for section in sections:
        company = section["company"]
        site = section["site"]
        for rec in section["records"]:
            alt = row % 2 == 0
            for col, column in enumerate(columns, start=1):
                if column["kind"] == "distractor":
                    cell_value = resolve_tabular_value(distractor_plan, column["distractor_field"], _heat_row_key(company, site, rec))
                else:
                    value = column["accessor"](company, site, rec)
                    cell_value = None if (column["blank_field"] and column["blank_field"] in blank_fields) else value
                _data_cell(ws.cell(row=row, column=col), cell_value, fmt=column["fmt"], alt=alt)
            ws.row_dimensions[row].height = 18
            row += 1


def generate_xlsx(
    config: dict,
    sections: list[dict],
    blank_fields: set[str] | None = None,
    split_by_company: bool = False,
    include_summary: bool = True,
    category: str = "heat",
) -> bytes:
    document_type = config["document"].get("type")
    if category == "electricity":
        if document_type == "smart_meter_data":
            return _generate_smart_meter_xlsx(config, sections)
        if document_type == "supplier_portal_data":
            return _generate_electricity_supplier_portal_xlsx(config, sections, include_summary=include_summary)
        raise NotImplementedError(f"XLSX generation is not supported for electricity document type '{document_type}'.")
    if document_type == "supplier_portal_data":
        return _generate_heat_supplier_portal_xlsx(
            config,
            sections,
            blank_fields=blank_fields,
            split_by_company=split_by_company,
            include_summary=include_summary,
        )
    raise NotImplementedError(f"XLSX generation is not supported for heat document type '{document_type}'.")
